# pip install openpyxl==3.0.7
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws1 = wb.create_sheet("시트0") # 가장 뒤에 시트 생성 (기본값)
ws2 = wb.create_sheet("시트1", 0) # 가장 앞에 시트 생성 
ws3 = wb.create_sheet("시트2", -1) # 끝에서 두 번째에 시트 생성

ws1.title='시트3'

print(wb.sheetnames)

wb.save('wb.xlsx')
